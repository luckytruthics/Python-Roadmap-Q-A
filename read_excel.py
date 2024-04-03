import openpyxl as xl
import os
from datetime import datetime,date



# Loading the workbook
workbook = xl.load_workbook(os.path.join(os.path.dirname(__file__),'prospect_sales1.xlsx'))


# Get the market name from the sheet
def get_market_name(sheet):
    for row in sheet:
        if row[0].value == 'Market:':
            return row[4].value
   


# Get all the requested columns from the sheet
def get_cols_index(sheet):
    col_index = {}
    for index,cols in enumerate(sheet.iter_cols(values_only=True)):
        if 'DATE' in cols:
            col_index['DATE'] = index

        elif 'COST' in cols:
            col_index['COST'] = index

        elif 'LEN' in cols:
            col_index['LEN'] = index

        elif 'EST (000)' in cols:
            col_index['EST (000)'] = index


    return col_index




def get_sheet_records(sheet):
    sheet_records = []
    col_index = get_cols_index(sheet)
    market_name =  get_market_name(sheet)
    for row in sheet.iter_rows(values_only=True):
        if not row[0] and row[18]:
            month_date = row[col_index['DATE']].split('/')
            if int(month_date[0]) > 12:
                date_var,month_var = month_date[0],month_date[1]
            else:
                month_var,date_var = month_date[0],month_date[1]

            date_obj = date(2019,int(month_var),int(date_var))
            sheet_records.append([date_obj,market_name,row[col_index['COST']],row[18],row[col_index['EST (000)']]])

    return sheet_records




# adding the header in the excel file

def set_header(sheet_name,headers):
    for col_index,col_header in enumerate(headers,start=1):
        sheet_name.cell(row=1,column = col_index).value = col_header




# Adding the data into sheets

write_wb = xl.Workbook()
# write_sheet = write_wb.active


def write_excel(): 
    headers = ['Date','Market','Cost','Len','EST (000)']
    for index,sheet in enumerate(workbook.worksheets):
        data = get_sheet_records(sheet)
        write_sheet = write_wb.create_sheet()
        write_sheet.title = f"Sheet {index+1}"
        set_header(write_sheet,headers)
        for row in data:
            write_sheet.append(row)
        


write_excel()

write_wb.save('revised_extracted_sales.xlsx')



